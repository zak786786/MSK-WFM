import pandas as pd
import requests
import openpyxl
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from fetch_auction_data import fetch_auction_data

start = time.time()

# Read data from excel
file_path = r'D:\Documents\Warframe.xlsx'
df = pd.read_excel(file_path, sheet_name='Weapons')
my_list = df.iloc[:, 0].tolist()

# Prepare weapon names
weapons = [s.lower().replace(" ", "_") for s in my_list]

# Load worksheet
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook['Weapons']

# Create session object for requests
session = requests.Session()

df1 = pd.DataFrame(weapons)
df1.columns = ['Weapons']
df1['Price'] = pd.Series([float('nan')] * len(df1))

# Use ThreadPoolExecutor to execute requests concurrently
futures = []
with ThreadPoolExecutor(max_workers=10) as executor:
    for weapon in weapons:
        # print(weapon)
        future = executor.submit(fetch_auction_data, weapon, session)
        # print(future.result)
        futures.append(future)
    # Get results as they complete
    for i, future in enumerate(as_completed(futures)):
        buyout_price, name = future.result()
        print(future.result())
        if buyout_price is not None:
            # worksheet.cell(row=i + 2, column=5, value=buyout_price)
            df1.loc[df1['Weapons'] == name, 'Price'] = buyout_price

# print(df1)
for i in range(len(weapons)):
    worksheet.cell(row=i + 2, column=5, value=df1.iloc[i, 1])

# Save workbook
workbook.save(file_path)

# Open the Excel file with the default application
os.startfile(file_path)

end = time.time()
print(end - start)

